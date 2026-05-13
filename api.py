"""
API мост между JS-интерфейсом и Python-логикой
JavaScript вызывает методы класса API через window.pywebview.api.method_name(...)
"""
import os
import sys
import json
import asyncio
import webview
from datetime import datetime
from typing import List, Dict, Any

from scrapers import search_all_sources
from classifier import classify_suppliers
from excel_export import export_to_excel


VAT_RATES = {
    "UZ": 12, "AZ": 18, "KZ": 12, "KG": 12, "TJ": 18, "TM": 15,
    "RU": 20, "TR": 20, "CN": 13, "DE": 19, "US": 0, "AE": 5,
    "GB": 20, "PL": 23, "GE": 18, "AM": 20,
}


class API:
    """Методы доступные из JavaScript"""

    def __init__(self):
        self._window = None

    def _set_window(self, window):
        self._window = window

    def ping(self) -> Dict[str, str]:
        """Проверка что Python отвечает"""
        return {"status": "ok", "version": "1.0.0"}

    def search_suppliers(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """
        Главный метод — поиск поставщиков.
        Получает {country, region, currency, items: [{name, param, unit, qty}, ...]}
        Возвращает результаты для всех позиций.
        """
        try:
            country = payload.get("country", "UZ")
            region = payload.get("region", "")
            currency = payload.get("currency", "UZS")
            items = payload.get("items", [])

            if not items:
                return {"error": "Список позиций пуст"}

            vat_rate = VAT_RATES.get(country, 12)
            rows = []

            # Запускаем event loop для асинхронных скраперов
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            try:
                for idx, item in enumerate(items, 1):
                    # Уведомление UI о прогрессе
                    self._notify_progress(idx, len(items), item.get('name', ''))

                    # Запрос для скрапинга
                    query = item.get("name", "")
                    if item.get("param"):
                        query += f" {item['param']}"

                    # Скрапим разные источники
                    listings = loop.run_until_complete(
                        search_all_sources(query, country=country, region=region)
                    )

                    # Классифицируем — передаём query для фильтрации мусора
                    suppliers = classify_suppliers(listings, item.get("qty", 1), query=query)

                    rows.append({
                        "item_num": idx,
                        "item_name": item.get("name", ""),
                        "item_param": item.get("param", ""),
                        "item_unit": item.get("unit", "шт"),
                        "item_qty": item.get("qty", 1),
                        "suppliers": [s.dict() for s in suppliers],
                    })
            finally:
                loop.close()

            # Считаем диапазоны итоговых сумм
            total_min = sum(
                min((s["total"] for s in r["suppliers"]), default=0)
                for r in rows
            )
            total_max = sum(
                max((s["total"] for s in r["suppliers"]), default=0)
                for r in rows
            )

            return {
                "country": country,
                "region": region,
                "currency": currency,
                "vat_rate": vat_rate,
                "rows": rows,
                "total_min": round(total_min, 2),
                "total_max": round(total_max, 2),
                "created_at": datetime.now().isoformat(),
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"error": f"Ошибка поиска: {str(e)}"}

    def _notify_progress(self, current: int, total: int, item_name: str):
        """Отправить уведомление о прогрессе в UI"""
        if self._window:
            try:
                self._window.evaluate_js(
                    f"window.onSearchProgress && window.onSearchProgress("
                    f"{current}, {total}, {json.dumps(item_name)})"
                )
            except Exception:
                pass

    def export_excel(self, result: Dict[str, Any], country_name: str, save_path: str = None) -> Dict[str, Any]:
        """Экспорт результатов в Excel"""
        try:
            if not save_path:
                # Открыть диалог сохранения
                if self._window:
                    files = self._window.create_file_dialog(
                        dialog_type=webview.SAVE_DIALOG,
                        save_filename=f"procurement_{country_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        file_types=("Excel files (*.xlsx)",),
                    )
                    if not files:
                        return {"cancelled": True}
                    save_path = files[0] if isinstance(files, (list, tuple)) else files

            export_to_excel(result, country_name, save_path)
            return {"ok": True, "path": save_path}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"error": str(e)}

    def import_excel(self) -> Dict[str, Any]:
        """Импорт ТМЦ из Excel-файла"""
        try:
            if not self._window:
                return {"error": "Окно недоступно"}

            files = self._window.create_file_dialog(
                dialog_type=webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("Excel files (*.xlsx;*.xls)",),
            )
            if not files:
                return {"cancelled": True}

            file_path = files[0] if isinstance(files, (list, tuple)) else files

            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            items = []
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:  # пропускаем заголовок
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip() if row[0] else ""
                param = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                unit = str(row[2]).strip() if len(row) > 2 and row[2] else "шт"
                qty_raw = row[3] if len(row) > 3 else None
                try:
                    qty = float(qty_raw) if qty_raw is not None else 1.0
                except (ValueError, TypeError):
                    qty = 1.0
                if name and qty > 0:
                    items.append({
                        "name": name, "param": param, "unit": unit, "qty": qty,
                    })
            wb.close()

            return {"ok": True, "items": items, "count": len(items)}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"error": str(e)}
